import openpyxl
def getData(sheetName):

    workbook = openpyxl.load_workbook("..//excel//testData.xlsx")
    sheet = workbook[sheetName]
    total_row = sheet.max_row
    total_cols = sheet.max_column
    mainList=[]
    for i in range(2,total_row+1):
        dataList=[]
        for j in range(1,total_cols+1):
            data=sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainList.insert(i,dataList)
    #print(mainList)
    return mainList